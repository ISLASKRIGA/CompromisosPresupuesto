import openpyxl
import json
import os

EXCEL_FILE = r'Pruebareporteejecutivo_validado_PCOM_CM_correcto.xlsx'

def to_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace('$', '').replace(',', '')
    if s in ('-', '$-', '$ -', '', 'N/A', 'None', '#VALUE!', '#N/A'):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def process_excel():
    print(f"Cargando {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Base de datos origen']

    records = []
    for r in range(3, ws.max_row + 1):
        row_id = ws.cell(r, 1).value
        cto = ws.cell(r, 6).value
        desc = ws.cell(r, 4).value
        if row_id is None and cto is None and desc is None:
            continue

        no_comp = str(ws.cell(r, 33).value or '').strip()
        if no_comp and not no_comp.startswith('Compromiso') and no_comp != 'SIN COMPROMISO' and no_comp.isdigit():
            no_comp = f'Compromiso #{no_comp}'
        elif not no_comp or no_comp == '0':
            no_comp = 'SIN FOLIO COMPROMISO'

        autorizado = to_float(ws.cell(r, 28).value)
        sicop_mod = to_float(ws.cell(r, 36).value)
        sicop_ejer = to_float(ws.cell(r, 37).value)
        dif_mod = to_float(ws.cell(r, 38).value)
        pagado_sicop = to_float(ws.cell(r, 40).value)
        pagado_almacen = to_float(ws.cell(r, 41).value)
        dif_pagado = to_float(ws.cell(r, 42).value)
        ejercer_sicop = to_float(ws.cell(r, 45).value)
        ejercer_est = to_float(ws.cell(r, 47).value or ws.cell(r, 44).value)
        dif_ejercer = to_float(ws.cell(r, 46).value)

        ptda_val = str(ws.cell(r, 18).value or '').strip()
        capitulo = str(ws.cell(r, 53).value or ('2000' if ptda_val.startswith('2') else '3000')).strip()

        comp_sicop = str(ws.cell(r, 35).value or 'NO').strip().upper()
        if comp_sicop not in ('SI', 'NO'):
            comp_sicop = 'SI' if sicop_mod > 0 else 'NO'

        obs = str(ws.cell(r, 50).value or '').strip()
        if not obs:
            if abs(dif_mod) < 0.01:
                obs = 'Conciliación por contrato y compromiso correcta'
            elif dif_mod > 0:
                obs = f'Sub-registro en Almacén INPER (Omisión de capturar ${dif_mod:,.2f} MXN en reporte)'
            else:
                obs = f'Sobre-registro en Almacén INPER (${abs(dif_mod):,.2f} MXN por encima de SICOP)'

        rec = {
            'id': int(row_id) if row_id is not None and str(row_id).isdigit() else (len(records) + 1),
            'fila_excel': f'Fila {r}',
            'area': str(ws.cell(r, 3).value or '').strip(),
            'descripcion': str(ws.cell(r, 4).value or '').strip(),
            'contrato': str(cto or '').strip(),
            'no_compromiso': no_comp,
            'anexo': str(ws.cell(r, 7).value or 'N/A').strip(),
            'proveedor': str(ws.cell(r, 8).value or '').strip(),
            'rfc': str(ws.cell(r, 9).value or '').strip(),
            'autorizado': autorizado,
            'comprometido_sicop': comp_sicop,
            'sicop_mod': sicop_mod,
            'sicop_ejer': sicop_ejer,
            'dif_mod': dif_mod,
            'pagado_sicop': pagado_sicop,
            'pagado_almacen': pagado_almacen,
            'dif_pagado': dif_pagado,
            'ejercer_sicop': ejercer_sicop,
            'ejercer_est': ejercer_est,
            'dif_ejercer': dif_ejercer,
            'prioridad': str(ws.cell(r, 34).value or 'ALTA').strip(),
            'capitulo': capitulo,
            'que_paso': obs,
            'partida': ptda_val
        }
        records.append(rec)

    # 1. Build Contracts Summary
    contract_map = {}
    for r in records:
        c_id = r['contrato']
        if not c_id:
            c_id = 'SIN CONTRATO'
        if c_id not in contract_map:
            contract_map[c_id] = {
                'contrato': c_id,
                'no_compromiso': r['no_compromiso'],
                'proveedor': r['proveedor'],
                'capitulo': r['capitulo'],
                'comprometido_sicop': r['comprometido_sicop'],
                'anexo': r['anexo'],
                'prioridad': r['prioridad'],
                'autorizado_sum': 0.0,
                'sicop_mod': 0.0,
                'almacen_sum': 0.0,
                'dif_mod_contrato': 0.0,
                'pagado_sicop_sum': 0.0,
                'pagado_almacen_sum': 0.0,
                'dif_pagado_contract': 0.0,
                'ejercer_sicop_sum': 0.0,
                'ejercer_est_sum': 0.0,
                'dif_ejercer_contract': 0.0,
                'partidas_count': 0,
                'monto_max': 0.0
            }
        c_obj = contract_map[c_id]
        c_obj['autorizado_sum'] += r['autorizado']
        c_obj['sicop_mod'] += r['sicop_mod']
        c_obj['almacen_sum'] += r['sicop_ejer']
        c_obj['dif_mod_contrato'] += r['dif_mod']
        c_obj['pagado_sicop_sum'] += r['pagado_sicop']
        c_obj['pagado_almacen_sum'] += r['pagado_almacen']
        c_obj['dif_pagado_contract'] += r['dif_pagado']
        c_obj['ejercer_sicop_sum'] += r['ejercer_sicop']
        c_obj['ejercer_est_sum'] += r['ejercer_est']
        c_obj['dif_ejercer_contract'] += r['dif_ejercer']
        c_obj['partidas_count'] += 1
        if r['no_compromiso'] and r['no_compromiso'] != 'SIN FOLIO COMPROMISO':
            c_obj['no_compromiso'] = r['no_compromiso']

    contracts = list(contract_map.values())

    # Calculate global totals
    total_autorizado = sum(c['autorizado_sum'] for c in contracts)
    total_sicop = sum(c['sicop_mod'] for c in contracts)
    total_almacen = sum(c['almacen_sum'] for c in contracts)
    total_dif = sum(r['dif_mod'] for r in records)
    diff_records = [r for r in records if abs(r['dif_mod']) > 0.01]

    # Calculate sobrantes & faltantes by contract
    sobrantes_list = []
    faltantes_list = []
    sobrante_total = 0.0
    faltante_total = 0.0

    for c in contracts:
        diff_suf = c['autorizado_sum'] - c['sicop_mod']
        c['sobrante_suficiencia'] = diff_suf
        if diff_suf > 0.01:
            sobrantes_list.append(c)
            sobrante_total += diff_suf
        elif diff_suf < -0.01:
            c['faltante_suficiencia'] = abs(diff_suf)
            faltantes_list.append(c)
            faltante_total += abs(diff_suf)

    sobrantes_list.sort(key=lambda x: x['sobrante_suficiencia'], reverse=True)
    faltantes_list.sort(key=lambda x: x.get('faltante_suficiencia', 0.0), reverse=True)

    summary = {
        'total_records': len(records),
        'total_contracts': len(contracts),
        'autorizado_total': total_autorizado,
        'sicop_mod_total': total_sicop,
        'almacen_reported_total': total_almacen,
        'dif_mod_total': total_dif,
        'errors_count': len(diff_records),
        'coverage_percentage': round((total_sicop / total_autorizado * 100), 2) if total_autorizado > 0 else 0,
        'sobrante_total': sobrante_total,
        'faltante_total': faltante_total,
        'saldo_neto': sobrante_total - faltante_total
    }

    # Save dashboard_data.json
    dashboard_data = {
        'records': records,
        'contracts': contracts,
        'summary': summary
    }
    with open('dashboard_data.json', 'w', encoding='utf-8') as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=2)
    print("dashboard_data.json actualizado con éxito.")

    # Save almacen_errors.json
    sub_reg = [r for r in diff_records if r['dif_mod'] > 0]
    sobre_reg = [r for r in diff_records if r['dif_mod'] < 0]
    omision_tot = [r for r in diff_records if r['sicop_ejer'] == 0 and r['sicop_mod'] > 0]

    almacen_errors = {
        'total_records': len(records),
        'diff_records_count': len(diff_records),
        'sicop_mod_sum': sum(r['sicop_mod'] for r in diff_records),
        'inper_mod_sum': sum(r['sicop_ejer'] for r in diff_records),
        'dif_mod_sum': sum(r['dif_mod'] for r in diff_records),
        'sub_registro_count': len(sub_reg),
        'sub_registro_sum': sum(r['dif_mod'] for r in sub_reg),
        'sobre_registro_count': len(sobre_reg),
        'sobre_registro_sum': sum(r['dif_mod'] for r in sobre_reg),
        'omision_total_count': len(omision_tot),
        'omision_total_sum': sum(r['dif_mod'] for r in omision_tot),
        'top_15_errors': sorted(diff_records, key=lambda x: abs(x['dif_mod']), reverse=True)[:15]
    }
    with open('almacen_errors.json', 'w', encoding='utf-8') as f:
        json.dump(almacen_errors, f, ensure_ascii=False, indent=2)
    print("almacen_errors.json actualizado con éxito.")

    # Save actionable_data.json
    # Build balanza_partidas
    ptda_map = {}
    for r in records:
        ptda = r['partida'] or 'OTRAS'
        if ptda not in ptda_map:
            ptda_map[ptda] = {
                'clave': ptda,
                'nombre': f"{ptda} - {r['descripcion'][:40]}",
                'capitulo': r['capitulo'],
                'contratos_count': 0,
                'autorizado': 0.0,
                'sicop': 0.0,
                'sobrante': 0.0,
                'sobrantes_count': 0,
                'faltante': 0.0,
                'faltantes_count': 0,
                'neto': 0.0,
                'estrategia': 'CONCILIAR SUFICIENCIA',
                'diagnostico': 'Revisión presupuestal por clave de partida',
                'accion_tipo': 'REVISIÓN'
            }
        p_obj = ptda_map[ptda]
        p_obj['autorizado'] += r['autorizado']
        p_obj['sicop'] += r['sicop_mod']
        p_obj['contratos_count'] += 1
        d_val = r['autorizado'] - r['sicop_mod']
        if d_val > 0.01:
            p_obj['sobrante'] += d_val
            p_obj['sobrantes_count'] += 1
        elif d_val < -0.01:
            p_obj['faltante'] += abs(d_val)
            p_obj['faltantes_count'] += 1

    for p in ptda_map.values():
        p['neto'] = p['sobrante'] - p['faltante']
        if p['neto'] > 0.01:
            p['estrategia'] = f"LIBERAR REMANENTE DE ${p['neto']/1e6:.2f} MDP: Disponible para reasignar"
            p['diagnostico'] = "Excedente de suficiencia sin compromiso exigible"
            p['accion_tipo'] = "LIBERAR"
        elif p['neto'] < -0.01:
            p['estrategia'] = f"AMPLIAR SUFICIENCIA EN ${abs(p['neto'])/1e6:.2f} MDP: Cobertura CM1"
            p['diagnostico'] = "Déficit por convenios modificatorios"
            p['accion_tipo'] = "COBERTURA"

    actionable_data = {
        'sobrantes': sobrantes_list[:52],
        'faltantes': faltantes_list[:43],
        'sobrante_total': sobrante_total,
        'faltante_total': faltante_total,
        'net_balance': sobrante_total - faltante_total,
        'balanza_partidas': list(ptda_map.values())
    }
    with open('actionable_data.json', 'w', encoding='utf-8') as f:
        json.dump(actionable_data, f, ensure_ascii=False, indent=2)
    print("actionable_data.json actualizado con éxito.")

    # Save analysis_breakdown.json
    by_partida = list(ptda_map.values())
    by_contrato = sorted(contracts, key=lambda x: x['autorizado_sum'], reverse=True)[:50]
    analysis_breakdown = {
        'by_partida': by_partida,
        'by_contrato': by_contrato,
        'top_inflado': sorted(contracts, key=lambda x: abs(x['dif_mod_contrato']), reverse=True)[:20]
    }
    with open('analysis_breakdown.json', 'w', encoding='utf-8') as f:
        json.dump(analysis_breakdown, f, ensure_ascii=False, indent=2)
    print("analysis_breakdown.json actualizado con éxito.")

if __name__ == '__main__':
    process_excel()
